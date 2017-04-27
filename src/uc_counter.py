from collections import defaultdict
from openpyxl import load_workbook
sole = defaultdict(int)
from openpyxl import Workbook
wb = load_workbook("../podatki/izbirci_correct.xlsx")
sr= wb["correcttxt"]
c=2
while c!=77747:
    sola=sr['B'+str(c)].value
    ucenci=sr['H'+str(c)].value
    sole[sola]+=ucenci
    c+=1

wb_new = Workbook()
dest="../nasi_podatki/statistika_otrok_pp.xlsx"
ws1= wb_new.active
ws1.title = "St. otrok na solo"
ws1['A1']="Šola"
ws1['B1']="Število otrok"
counter=2
for key,value in sole.items():
    ws1['A'+str(counter)]=key
    ws1['B'+str(counter)]=value
    print(key,value)
    counter+=1
wb_new.save(filename=dest)